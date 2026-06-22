import os
import tempfile
import unittest

from openpyxl import load_workbook

from src.fake_news.reporting.report_card import ExperimentRecord, ModelReport

_METRICS_BASELINE = {'accuracy': 0.2, 'macro_f1': 0.1, 'macro_precision': 0.1, 'macro_recall': 0.2}
_METRICS_IMPROVED = {'accuracy': 0.4, 'macro_f1': 0.2, 'macro_precision': 0.3, 'macro_recall': 0.3}


def _baseline():
    return ExperimentRecord('Baseline', {'strategy': 'most_frequent'}, _METRICS_BASELINE,
                            'baseline')


def _improved():
    return ExperimentRecord('LSTM', {'hidden_size': 128}, _METRICS_IMPROVED, 'better')


class TestModelReportInit(unittest.TestCase):

    def test_when_main_metric_invalid_then_value_error_is_raised(self):
        with self.assertRaises(ValueError):
            ModelReport(main_metric='nonsense')


class TestModelReportBaseline(unittest.TestCase):

    def test_when_no_records_then_value_error_is_raised(self):
        with self.assertRaises(ValueError):
            _ = ModelReport().baseline

    def test_when_records_added_then_baseline_is_first(self):
        report = ModelReport().add(_baseline()).add(_improved())
        self.assertEqual(report.baseline.name, 'Baseline')


class TestModelReportPercentChange(unittest.TestCase):

    def test_when_metric_improves_then_change_is_positive(self):
        report = ModelReport().add(_baseline()).add(_improved())
        self.assertAlmostEqual(report.percent_change(_improved(), 'macro_f1'), 100.0)

    def test_when_baseline_metric_zero_then_change_is_zero(self):
        zero = ExperimentRecord('B', {}, {
            'accuracy': 0.0,
            'macro_f1': 0.0,
            'macro_precision': 0.0,
            'macro_recall': 0.0
        })
        report = ModelReport().add(zero).add(_improved())
        self.assertEqual(report.percent_change(_improved(), 'macro_f1'), 0.0)


class TestModelReportBestRecordIndex(unittest.TestCase):

    def test_when_no_records_then_value_error_is_raised(self):
        with self.assertRaises(ValueError):
            ModelReport().best_record_index()

    def test_when_records_added_then_returns_index_of_highest_main_metric(self):
        report = ModelReport().add(_baseline()).add(_improved())
        self.assertEqual(report.best_record_index(), 1)


class TestModelReportSave(unittest.TestCase):

    def setUp(self):
        self.directory = tempfile.mkdtemp()
        self.path = os.path.join(self.directory, 'report.xlsx')

    def test_when_empty_then_value_error_is_raised(self):
        with self.assertRaises(ValueError):
            ModelReport().save(self.path)

    def test_when_saved_then_workbook_has_expected_sheets(self):
        report = ModelReport().add(_baseline()).add(_improved())
        diagram = os.path.join(self.directory, 'missing.png')
        report.save(self.path,
                    diagram_paths=[diagram],
                    examples=[('some text', 'true', 'true'), ('other', 'false', 'true')])
        workbook = load_workbook(self.path)
        self.assertIn('Model Report', workbook.sheetnames)
        self.assertIn('Diagrams', workbook.sheetnames)
        self.assertIn('Best Model Examples', workbook.sheetnames)

    def test_when_saved_then_baseline_metric_cell_marked_baseline(self):
        report = ModelReport().add(_baseline()).add(_improved())
        report.save(self.path)
        sheet = load_workbook(self.path)['Model Report']
        flattened = [cell.value for row in sheet.iter_rows() for cell in row]
        self.assertTrue(any('baseline' in str(value) for value in flattened))


if __name__ == '__main__':
    unittest.main()
