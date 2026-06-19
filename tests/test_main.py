import os
import tempfile
import unittest

import pandas as pd

from src.fake_news.config import DataConfig, ExperimentConfig, LSTMConfig
from src.fake_news.main import (_collect_examples, load_dataframe, resolve_output_paths, run)


def _write_isot(directory):
    fake_path = os.path.join(directory, 'Fake.csv')
    true_path = os.path.join(directory, 'True.csv')
    pd.DataFrame({'title': ['a'], 'text': ['x']}).to_csv(fake_path, index=False)
    pd.DataFrame({'title': ['b'], 'text': ['y']}).to_csv(true_path, index=False)
    return fake_path, true_path


def _tiny_config(directory: str) -> ExperimentConfig:
    return ExperimentConfig(
        data=DataConfig(fake_path=os.path.join(directory, 'missing_fake.csv'),
                        true_path=os.path.join(directory, 'missing_true.csv'),
                        min_token_frequency=1,
                        max_sequence_length=20),
        lstm=LSTMConfig(embedding_dim=8, hidden_size=8, epochs=1, batch_size=32),
        report_path=os.path.join(directory, 'report.xlsx'),
        figures_dir=os.path.join(directory, 'figures'),
    )


class TestLoadDataframe(unittest.TestCase):

    def test_when_auto_and_isot_missing_then_falls_back_to_synthetic(self):
        config = _tiny_config(tempfile.mkdtemp())
        frame, source = load_dataframe(config)
        self.assertEqual(source, 'synthetic')
        self.assertGreater(len(frame), 0)

    def test_when_auto_and_isot_present_then_loads_isot(self):
        directory = tempfile.mkdtemp()
        fake_path, true_path = _write_isot(directory)
        config = ExperimentConfig(data=DataConfig(fake_path=fake_path, true_path=true_path))
        _, source = load_dataframe(config)
        self.assertEqual(source, 'ISOT')

    def test_when_source_is_synthetic_then_ignores_existing_csvs(self):
        directory = tempfile.mkdtemp()
        fake_path, true_path = _write_isot(directory)
        config = ExperimentConfig(
            data=DataConfig(fake_path=fake_path, true_path=true_path, data_source='synthetic'))
        _, source = load_dataframe(config)
        self.assertEqual(source, 'synthetic')

    def test_when_source_is_real_and_present_then_loads_isot(self):
        directory = tempfile.mkdtemp()
        fake_path, true_path = _write_isot(directory)
        config = ExperimentConfig(
            data=DataConfig(fake_path=fake_path, true_path=true_path, data_source='real'))
        _, source = load_dataframe(config)
        self.assertEqual(source, 'ISOT')

    def test_when_source_is_real_and_missing_then_file_not_found_is_raised(self):
        config = _tiny_config(tempfile.mkdtemp())
        config.data.data_source = 'real'
        with self.assertRaises(FileNotFoundError):
            load_dataframe(config)

    def test_when_source_is_invalid_then_value_error_is_raised(self):
        config = _tiny_config(tempfile.mkdtemp())
        config.data.data_source = 'nonsense'
        with self.assertRaises(ValueError):
            load_dataframe(config)


class TestResolveOutputPaths(unittest.TestCase):

    def test_when_called_then_source_is_inserted_into_paths(self):
        config = ExperimentConfig(report_path='reports/model_report.xlsx',
                                  figures_dir='reports/figures')
        report_path, figures_dir = resolve_output_paths(config, 'ISOT')
        self.assertEqual(report_path, 'reports/model_report_isot.xlsx')
        self.assertEqual(figures_dir, os.path.join('reports/figures', 'isot'))


class TestCollectExamples(unittest.TestCase):

    def test_when_called_then_mixes_correct_and_incorrect(self):
        texts = ['a', 'b', 'c', 'd']
        references = [0, 1, 0, 1]
        predictions = [0, 1, 1, 0]
        examples = _collect_examples(texts, references, predictions, limit=4)
        self.assertEqual(len(examples), 4)


class TestRun(unittest.TestCase):

    def test_when_run_then_per_source_report_and_figures_are_produced(self):
        directory = tempfile.mkdtemp()
        config = _tiny_config(directory)
        exit_code = run(config)
        report_path, figures_dir = resolve_output_paths(config, 'synthetic')
        self.assertEqual(exit_code, 0)
        self.assertTrue(os.path.exists(report_path))
        self.assertTrue(os.path.exists(os.path.join(figures_dir, 'class_distribution.png')))

    def test_when_run_then_report_records_the_data_source(self):
        from openpyxl import load_workbook
        directory = tempfile.mkdtemp()
        config = _tiny_config(directory)
        run(config)
        report_path, _ = resolve_output_paths(config, 'synthetic')
        sheet = load_workbook(report_path)['Model Report']
        flattened = [str(cell.value) for row in sheet.iter_rows() for cell in row]
        self.assertTrue(any('synthetic' in value for value in flattened))


if __name__ == '__main__':
    unittest.main()
