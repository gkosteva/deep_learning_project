import os
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

METRIC_KEYS = ('accuracy', 'f1', 'recall')
METRIC_TITLES = {'accuracy': 'Accuracy', 'f1': 'F1', 'recall': 'Recall'}

_HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
_HEADER_FONT = Font(color='FFFFFF', bold=True)
_BASELINE_FILL = PatternFill('solid', fgColor='FCE4D6')
_BEST_ROW_FILL = PatternFill('solid', fgColor='C6EFCE')
_BEST_METRIC_FILL = PatternFill('solid', fgColor='FFEB9C')
_TITLE_FONT = Font(size=14, bold=True)
_BEST_FONT = Font(bold=True, color='006100')
_THIN = Side(style='thin', color='BFBFBF')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


@dataclass
class ExperimentRecord:
    name: str
    hyperparameters: Dict[str, object] = field(default_factory=dict)
    metrics: Dict[str, float] = field(default_factory=dict)
    comment: str = ''


class ModelReport:

    def __init__(self, title: str = 'Model Report - Fake News Detection', main_metric: str = 'f1'):
        if main_metric not in METRIC_KEYS:
            raise ValueError(f'main_metric must be one of {METRIC_KEYS}')
        self.title = title
        self.main_metric = main_metric
        self.records: List[ExperimentRecord] = []

    def add(self, record: ExperimentRecord) -> 'ModelReport':
        self.records.append(record)
        return self

    @property
    def baseline(self) -> ExperimentRecord:
        if not self.records:
            raise ValueError('report has no records; add the baseline first')
        return self.records[0]

    def percent_change(self, record: ExperimentRecord, metric: str) -> float:
        baseline_value = self.baseline.metrics.get(metric, 0.0)
        value = record.metrics.get(metric, 0.0)
        if baseline_value == 0:
            return 0.0
        return (value - baseline_value) / baseline_value * 100.0

    def best_record_index(self) -> int:
        if not self.records:
            raise ValueError('report has no records')
        return max(
            range(len(self.records)),
            key=lambda i: self.records[i].metrics.get(self.main_metric, 0.0),
        )

    def _hyperparameter_columns(self) -> List[str]:
        columns: List[str] = []
        for record in self.records:
            for key in record.hyperparameters:
                if key not in columns:
                    columns.append(key)
        return columns

    def _format_metric(self, record: ExperimentRecord, metric: str, is_baseline: bool) -> str:
        value = record.metrics.get(metric, 0.0)
        if is_baseline:
            return f'{value:.4f} (baseline)'
        change = self.percent_change(record, metric)
        sign = '+' if change >= 0 else ''
        return f'{value:.4f} ({sign}{change:.1f}%)'

    def save(
        self,
        path: str,
        diagram_paths: Optional[Sequence[str]] = None,
        examples: Optional[Sequence[Tuple[str, int, int]]] = None,
    ) -> str:
        if not self.records:
            raise ValueError('cannot save an empty report')
        directory = os.path.dirname(path)
        if directory:
            os.makedirs(directory, exist_ok=True)

        workbook = Workbook()
        self._write_main_sheet(workbook.active)
        self._write_diagrams_sheet(workbook.create_sheet('Diagrams'), diagram_paths)
        if examples is not None:
            self._write_examples_sheet(workbook.create_sheet('Best Model Examples'), examples)
        workbook.save(path)
        return path

    def _write_main_sheet(self, sheet) -> None:
        sheet.title = 'Model Report'
        hyper_columns = self._hyperparameter_columns()
        headers = (['Model'] + hyper_columns + [METRIC_TITLES[m]
                                                for m in METRIC_KEYS] + ['Comments'])
        best_index = self.best_record_index()
        best_record = self.records[best_index]

        sheet.cell(row=1, column=1, value=self.title).font = _TITLE_FONT
        best_cell = sheet.cell(
            row=2,
            column=1,
            value=(f'Best model: {best_record.name} - highest '
                   f'{METRIC_TITLES[self.main_metric]} '
                   f'({best_record.metrics.get(self.main_metric, 0.0):.4f})'),
        )
        best_cell.font = _BEST_FONT
        best_cell.fill = _BEST_METRIC_FILL

        header_row = 4
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=header_row, column=column_index, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = _BORDER

        metric_best_value = {
            metric: max(r.metrics.get(metric, 0.0)
                        for r in self.records)
            for metric in METRIC_KEYS
        }

        for offset, record in enumerate(self.records):
            row = header_row + 1 + offset
            is_baseline = offset == 0
            is_best = offset == best_index
            values = [record.name]
            values += [record.hyperparameters.get(col, '') for col in hyper_columns]
            metric_start = 1 + len(hyper_columns) + 1
            values += [self._format_metric(record, m, is_baseline) for m in METRIC_KEYS]
            values.append(record.comment)

            for column_index, value in enumerate(values, start=1):
                cell = sheet.cell(row=row, column=column_index, value=value)
                cell.border = _BORDER
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if is_best:
                    cell.fill = _BEST_ROW_FILL
                    cell.font = Font(bold=True)
                elif is_baseline:
                    cell.fill = _BASELINE_FILL

            for metric_offset, metric in enumerate(METRIC_KEYS):
                if record.metrics.get(metric, 0.0) == metric_best_value[metric]:
                    column_index = metric_start + metric_offset
                    sheet.cell(row=row, column=column_index).fill = _BEST_METRIC_FILL

        self._autosize_columns(sheet, headers)

    @staticmethod
    def _autosize_columns(sheet, headers: List[str]) -> None:
        for column_index, header in enumerate(headers, start=1):
            letter = get_column_letter(column_index)
            width = max(12, min(40, len(str(header)) + 4))
            if header == 'Comments':
                width = 45
            sheet.column_dimensions[letter].width = width

    @staticmethod
    def _write_diagrams_sheet(sheet, diagram_paths: Optional[Sequence[str]]) -> None:
        sheet.cell(row=1, column=1, value='Training diagrams').font = _TITLE_FONT
        if not diagram_paths:
            return
        anchor_row = 3
        for diagram_path in diagram_paths:
            if diagram_path and os.path.exists(diagram_path):
                image = XlsxImage(diagram_path)
                sheet.add_image(image, f'A{anchor_row}')
                anchor_row += 20

    @staticmethod
    def _write_examples_sheet(sheet, examples: Sequence[Tuple[str, int, int]]) -> None:
        sheet.cell(row=1, column=1, value='Best model - example predictions') \
            .font = _TITLE_FONT
        headers = ['Text (truncated)', 'True label', 'Predicted', 'Correct?']
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=column_index, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        for offset, (text, true_label, predicted) in enumerate(examples):
            row = 4 + offset
            correct = true_label == predicted
            sheet.cell(row=row, column=1, value=text[:120])
            sheet.cell(row=row, column=2, value=true_label)
            sheet.cell(row=row, column=3, value=predicted)
            cell = sheet.cell(row=row, column=4, value='yes' if correct else 'no')
            cell.fill = _BEST_ROW_FILL if correct else _BASELINE_FILL
        sheet.column_dimensions['A'].width = 70
